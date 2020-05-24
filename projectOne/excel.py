import openpyxl as xl
from openpyxl.chart import BarChart,Reference

wb = xl.load_workbook("transactions.xlsx")
sheet = wb["Sheet1"]
cell = sheet["a1"]
#Another way to read cell
cell = sheet.cell(1,1)
#print(cell.value)
noOfRows = sheet.max_row


for rows in range(2, noOfRows+1):
    cells = sheet.cell(rows,3)
    corrected_price = cells.value*0.9
    corrected_price_cell = sheet.cell(rows,4)
    corrected_price_cell.value = corrected_price

values = Reference(sheet, min_row = 2, max_row=sheet.max_row,min_col=4,max_col=4)
chart = BarChart()
chart.add_data(values)
sheet.add_chart(chart,'e2')
wb.save("transactions2.xlsx")