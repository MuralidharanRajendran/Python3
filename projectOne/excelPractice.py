import openpyxl as xl

wb = xl.load_workbook("transactions.xlsx")
sheet = wb["Sheet1"]
cell = sheet["a1"]
cell = sheet.cell(1,1)
print(cell.value)

noOfRows = sheet.max_row
print(noOfRows)

for rows in range(2,noOfRows+1):
    cell = sheet.cell(rows,3)
    print(cell.value)
    corrected_price = cell.value*0.9
    corrected_price_cell = sheet.cell(rows,4)
    corrected_price_cell.value = corrected_price
    tobeat = sheet.cell(1,4)
    tobeat.value = "new_update"
wb.save("transactions2.xlsx")